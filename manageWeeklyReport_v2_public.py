import openpyxl as pyxl
import requests
import json
import shutil
import pyfiglet
import sys
import os
import configparser
from datetime import datetime, timedelta

#使用前请config
config = configparser.ConfigParser()
#script_dir = os.path.dirname(sys.executable) 打包
script_dir = os.getcwd() #调试
config_path = script_dir + '\\config.txt'
config.read(config_path,encoding='utf-8')
config.default_section = configparser.DEFAULTSECT


parent_path = config.get('DEFAULT', 'parent_path')
parent_path = eval(parent_path)
template_file_name = config.get('DEFAULT', 'template_file_name')
template_file_name = eval(template_file_name)
name = config.get('DEFAULT', 'name')
name = eval(name)
api_key = config.get('DEFAULT', 'api_key')
api_key = eval(api_key)
api_base = config.get('DEFAULT', 'api_base')
api_base = eval(api_base)
gpt_model = config.get('DEFAULT', 'gpt_model')
gpt_model = eval(gpt_model)


def mon_fri_str():
    date = datetime.now() #今天日期的date对象
    weekday = date.weekday() #今天是周内的第几天(0-6)
    this_monday = date - timedelta(days=weekday) #本周一的date对象
    this_friday = this_monday + timedelta(days=4) #本周五的date对象
    #格式化字符串
    str_mon = this_monday.strftime('%m%d')
    str_fri = this_friday.strftime('%m%d')
    return str_mon,str_fri


#新建本周的周报文件
def create_file():
    str_mon,str_fri = mon_fri_str()
    template_path = parent_path + "\\" + template_file_name
    new_file_path = parent_path + "\\" + name + ' ' +f'{str_mon}-{str_fri}.xlsx'
    shutil.copy(template_path,new_file_path)
    return new_file_path

def create_initiate_file():
    #加载excel
    path = create_file()
    excel = pyxl.load_workbook(path)
    sheet = excel["Sheet1"]
    #初始化日期
    date = datetime.now() #今天日期的date对象
    weekday = date.weekday() #今天是周内的第几天(0-6)
    #获取周一到周五的对象
    this_week = [0]*5
    this_week[0] = date - timedelta(days=weekday)
    for i in range(1,5):
        this_week[i] = this_week[0] + timedelta(days=i)
    #修改年份
    for i in range(2,7):
        sheet[f'A{i}'] = this_week[i-2].strftime('%Y')+'年'
        sheet[f'B{i}'] = this_week[i-2].strftime('%m')+'月'
        sheet[f'C{i}'] = this_week[i-2].strftime('%d')+'日'
    excel.save(path)
    return path

def main_page():

    while True:
        title = """
        
.____     .__                
|    |    |__|___  __________
|    |    |  |\  \/ /\___   /
|    |___ |  | \   /  /    / 
|_______ ||__|  \_/  /_____ |

        """
        print(title)
        user_choice = input('>>>欢迎来到livz周报管理系统，请输入数字：\n[1]新建并初始化本周周报文件 [2] 写入周报内容 [3]生成本周周报文字版 [4]退出： ')
        try:
            user_choice = int(user_choice) # avoid invalid input
        except ValueError:
            error()
        if user_choice == 1:
            try:
                path = create_initiate_file()
                print(f'\n>>> 新建并初始化本周周报文件: 完成辣！\n 文件路径为：{path}')

            except Exception as e:
                error()

        elif user_choice == 2:
            print('\n')
            input_day = input('>>> 请问要输入周几的内容？[请输入数字1-5 下一步计划：6] ：')
            try:
                input_day = int(input_day)
            except ValueError:
                error()

            if input_day in [1,2,3,4,5]:
                week_list = ["一","二","三","四","五"]
                print(f"\n请在下方输入周{week_list[input_day-1]}的周报内容 [输入单行END结束输入]：")
                lines = []
                while True: 
                    line = input('[input]>>> ')
                    if line == 'END':
                        break
                    lines.append(line)

                try:
                    insert_context(lines,input_day-1)
                    print('\n>>> 写入周报内容：成功！')
                except Exception as e:
                    error()

            elif input_day == 6:
                print(f"\n请在下方输入下一步计划的内容 [输入单行END结束输入]：")
                lines = []
                while True: 
                    line = input('[input]>>> ')
                    if line == 'END':
                        break
                    lines.append(line)
                try:
                    print(lines)
                    insert_NextPlan(lines)
                    print('\n>>> 写入下一步计划内容：成功！')
                except Exception as e:
                    error()
                    print("注意：写入过程中如果文件被占用会导致错误")
            else:
                error()




        elif user_choice == 3:
            try:
                export_txt()
            except Exception as e:
                error()
                print(">>>导出错误")

        elif user_choice == 4:
            print("\n\n------------------------------------\n掰掰~~~~~~~~~~\n　 ∧＿∧\n （｡･ω･｡)つ━☆・*。\n ⊂　　 ノ 　　　・゜+.\n　 しーＪ　　　°。+ *´¨)\n　　　       　　.· ´¸.·*´¨) ¸.·*¨)\n 　　　　　　     　(¸.·´ (¸.·’*\")\"\n------------------------------------\n\n\n")
            sys.exit()


        else:
            error()


     
def error():
    print("\n>>>ERROR:出错啦，请检查输入")

def insert_context(lines,day):
    str_mon,str_fri = mon_fri_str()
    path = parent_path + "\\" + name + ' ' +f'{str_mon}-{str_fri}.xlsx'
    excel = pyxl.load_workbook(path)
    sheet = excel["Sheet1"]
    index = day+2
    content = ''
    for i in range(0,len(lines)):
        content = content + lines[i] + "\n"

    sheet[f'E{index}']= content
    excel.save(path)

def insert_NextPlan(lines):
    str_mon,str_fri = mon_fri_str()
    path = parent_path + "\\" + name + ' ' +f'{str_mon}-{str_fri}.xlsx'
    excel = pyxl.load_workbook(path)
    sheet = excel["Sheet1"]
    content = ''
    for i in range(0,len(lines)):
        content = content + lines[i] + "\n"
    print(content)
    sheet['F2'] = content
    excel.save(path)

def export_txt():
    print("\n>>>开始导出文字版.....")
    str_mon,str_fri = mon_fri_str()
    path = parent_path + "\\" + name + ' ' +f'{str_mon}-{str_fri}.xlsx'
    W_List = ["周一", "周二", "周三", "周四", "周五"]
    cell = []
    excel = pyxl.load_workbook(path)
    sheet = excel["Sheet1"]
    for i in range (2,7):
        cell_value = sheet["E"+str(i)].value
        cell.append(cell_value)

    cell_Next = sheet["F2"].value

    txt_path = parent_path + "\\" + name + ' ' +f'{str_mon}-{str_fri}.txt'
    txt = open(txt_path, 'w')
    for i in range(0, 5):
        txt.write(f"{W_List[i]}\n{cell[i]}\n\n")
    txt.close()

    with open(txt_path, 'r') as file:
        content_1 = file.read()

    print(">>>接入gpt中，请等待gpt响应........")

    content = gpt_generate(content_1)

    with open(txt_path, 'a') as file:  
        file.write("本周工作总结\n" + content + "\n")
        file.write(f"\n下一步计划 \n {cell_Next}")
    print(">>>导出周报文字版：完成！")

def gpt_generate(content):
    if api_key == 'None':
        print(">>>没有api key,不支持接入gpt，总结部分即将填入空字段")
        content = ''
        return content
    url = api_base
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json",
    }
    data = {
        "model": gpt_model,
        "messages": [
            {
                "role": "user",
                "content": f"这是我的本周工作内容：{content} 请列出三条工作总结，每一条前面标上序号，与我的格式相同。注意，请尽量简略。"
            }
        ]
    }

    response = requests.post(api_base, headers=headers, json=data)
    response_json = response.json()  # 将响应文本转换为JSON
    # 从响应JSON中提取content字段的值
    content = response_json["choices"][0]["message"]["content"]
    return content


main_page()

