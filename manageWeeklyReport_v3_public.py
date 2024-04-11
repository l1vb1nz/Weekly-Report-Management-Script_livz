import openpyxl as pyxl
import requests
import shutil
import pyfiglet
import sys
from datetime import datetime, timedelta

#使用前请config
parent_path = r'C:' #此处输入模板所在目录
template_file_name = '周报模版.xlsx' #此处输入模板文件名
name = '' #此处输入你的名字，生成的文件名会包含你的名字
api_key = '' #此处请填入api key，若无请填None
api_base = "" #此处请填入您的api接口地址，若无请忽视
gpt_model = 'gpt-3.5-turbo' #此处填入您想要使用的模型，若无请忽视


user_day = [0]*10
mon_fri = [0]*10
path = [0]*10


def mon_fri_str(n):
    num = n
    if num == 0:
        date = datetime.now() #今天日期的date对象
        weekday = date.weekday() #今天是周内的第几天(0-6)
        this_monday = date - timedelta(days=weekday) #本周一的date对象
        this_friday = this_monday + timedelta(days=4) #本周五的date对象
        #格式化字符串
        str_mon = this_monday.strftime('%m%d')
        str_fri = this_friday.strftime('%m%d')
        return str_mon,str_fri

    elif num == 1:
        print('\n>>>嘟嘟嘟......时光机启动............')
        year_input = input('>>>请问您想要穿梭到何年？: ')
        mon_input = input('>>>请问您想要穿梭到何月？[首位0忽略]: ')
        date_input = input('>>>请问您想要穿梭到何日？[首位0忽略]: ')
        try:
            year_input = int(year_input) # avoid invalid input
            mon_input = int(mon_input)
            date_input = int(date_input)
            date = datetime(year_input,mon_input,date_input)
            user_day[0] = year_input
            user_day[1] = mon_input
            user_day[2] = date_input
            weekday = date.weekday() #今天是周内的第几天(0-6)
            this_monday = date - timedelta(days=weekday) #本周一的date对象
            this_friday = this_monday + timedelta(days=4) #本周五的date对象
            #格式化字符串
            str_mon = this_monday.strftime('%m%d')
            str_fri = this_friday.strftime('%m%d')
            mon_fri[0] = str_mon
            mon_fri[1] = str_fri            
            return str_mon,str_fri
        except ValueError:
            error()
            return

#新建本周的周报文件
def create_file(n):
    num = n
    str_mon,str_fri = mon_fri_str(num)
    template_path = parent_path + "\\" + template_file_name
    new_file_path = parent_path + "\\" + name + ' ' +f'{str_mon}-{str_fri}.xlsx'
    path[0] = new_file_path
    shutil.copy(template_path,new_file_path)
    return new_file_path

def create_initiate_file(n):
    #加载excel
    num = n
    path = create_file(num)
    excel = pyxl.load_workbook(path)
    sheet = excel["Sheet1"]
    #初始化日期
    if num == 0:
        date = datetime.now() #今天日期的date对象        
    elif num == 1:
        date = datetime(user_day[0],user_day[1],user_day[2])
    weekday = date.weekday() #今天是周内的第几天(0-6)
    #获取周一到周五的对象
    this_week = [0]*5
    this_week[0] = date - timedelta(days=weekday)
    for i in range(1,5):
        this_week[i] = this_week[0] + timedelta(days=i)
    #修改年份
    for i in range(2,7):
        sheet[f'A{i}'] = this_week[i-2].strftime('%Y')+'年'
        sheet[f'B{i}'] = this_week[i-2].strftime('%m')+'月'
        sheet[f'C{i}'] = this_week[i-2].strftime('%d')+'日'
    excel.save(path)
    return path

def main_page():
    global path

    while True:
        title = pyfiglet.figlet_format('Livz', font='smslant')
        print(title)
        user_choice = input('>>>欢迎来到livz周报管理系统，请输入数字：\n[1]新建并初始化本周周报文件 [2]时间回溯新建并初始化 [3] 写入周报内容 [4]生成周报文字版 [5]退出： ')
        try:
            user_choice = int(user_choice) # avoid invalid input
        except ValueError:
            error()
        if user_choice == 1:
            try:
                create_initiate_file(0)
                print(f'\n>>> 新建并初始化本周周报文件: 完成辣！\n 文件路径为：{path[0]}')

            except Exception as e:
                error()

        elif user_choice == 3:
            print('\n')

            if path[0] == 0:
                str_mon,str_fri = mon_fri_str(0)
                path[0] = parent_path + "\\" + name + ' ' +f'{str_mon}-{str_fri}.xlsx'
            print(f'>>>当前路径为：{path[0]}')

            input_day = input('>>> 请问要输入周几的内容？[请输入数字1-5 下一步计划：6] [重置时间与路径：7]：')
            try:
                input_day = int(input_day)
            except ValueError:
                error()

            if input_day in [1,2,3,4,5]:
                week_list = ["一","二","三","四","五"]
                print(f"\n请在下方输入周{week_list[input_day-1]}的周报内容 [输入单行END结束输入]：")
                lines = []
                while True: 
                    line = input('[input]>>> ')
                    if line == 'END':
                        break
                    lines.append(line)

                try:
                    insert_context(lines,input_day-1,path[0])
                    print('\n>>> 写入周报内容：成功！')
                except Exception as e:
                    error()
                    print("注意：写入过程中如果文件被占用会导致错误")

            elif input_day == 6:
                print(f"\n请在下方输入下一步计划的内容 [输入单行END结束输入]：")
                lines = []
                while True: 
                    line = input('[input]>>> ')
                    if line == 'END':
                        break
                    lines.append(line)
                try:
                    print(lines)
                    insert_NextPlan(lines,path[0])
                    print('\n>>> 写入下一步计划内容：成功！')
                except Exception as e:
                    error()
                    print("注意：写入过程中如果文件被占用会导致错误")

            elif input_day == 7:
                str_mon,str_fri =mon_fri_str(1)
                path[0] = parent_path + "\\" + name + ' ' +f'{str_mon}-{str_fri}.xlsx'
                print(f'>>>路径：{path[0]}')


                         
            else:
                error()




        elif user_choice == 4:
            try:
                export_txt()
            except Exception as e:
                error()
                print(">>>导出错误")

        elif user_choice == 5:
            print("\n\n------------------------------------\n掰掰~~~~~~~~~~\n　 ∧＿∧\n （｡･ω･｡)つ━☆・*。\n ⊂　　 ノ 　　　・゜+.\n　 しーＪ　　　°。+ *´¨)\n　　　       　　.· ´¸.·*´¨) ¸.·*¨)\n 　　　　　　     　(¸.·´ (¸.·’*\")\"\n------------------------------------\n\n\n")
            sys.exit()

        elif user_choice == 2:
            try:
                create_initiate_file(1)
                print(f'\n>>> 新建并初始化时间回溯版周报文件: 完成辣！\n 文件路径为：{path[0]}')

            except Exception as e:
                error()            


        else:
            error()


     
def error():
    print("\n>>>ERROR:出错啦，请检查输入")

def insert_context(lines,day,path):
    excel = pyxl.load_workbook(path)
    sheet = excel["Sheet1"]
    index = day+2
    content = ''
    for i in range(0,len(lines)):
        content = content + lines[i] + "\n"

    sheet[f'E{index}']= content
    excel.save(path)

def insert_NextPlan(lines,path):
    excel = pyxl.load_workbook(path)
    sheet = excel["Sheet1"]
    content = ''
    for i in range(0,len(lines)):
        content = content + lines[i] + "\n"
    print(content)
    sheet['F2'] = content
    excel.save(path)

def export_txt():
    global path
    if path[0] == 0:
        str_mon,str_fri = mon_fri_str(0)
        path[0] = parent_path + "\\" + name + ' ' +f'{str_mon}-{str_fri}.xlsx'
    print(f'>>>当前路径为：{path[0]}')
    user_YN = input(f">>>是否是您想要的路径？[Y/N]： ")

    if user_YN in ["Y","y"]:
        pass

    elif user_YN in ["N","n"]:
        user_YN_1 = input(">>>请选择：[Y]本周 [N]其他: ")
        if user_YN_1 in ["Y","y"]:
            str_mon,str_fri = mon_fri_str(0)
            path[0] = parent_path + "\\" + name + ' ' +f'{str_mon}-{str_fri}.xlsx'
        elif user_YN_1 in ["N","n"]:
            str_mon,str_fri =mon_fri_str(1)
            path[0] = parent_path + "\\" + name + ' ' +f'{str_mon}-{str_fri}.xlsx'
        else:
            error()
    else:
        error()

    print(">>>开始导出文字版.....")
    W_List = ["周一", "周二", "周三", "周四", "周五"]
    cell = []
    excel = pyxl.load_workbook(path[0])
    sheet = excel["Sheet1"]
    for i in range (2,7):
        cell_value = sheet["E"+str(i)].value
        cell.append(cell_value)



    cell_Next = sheet["F2"].value

    txt_path = parent_path + "\\" + name + ' ' +f'{str_mon}-{str_fri}.txt'  
    txt = open(txt_path, 'w')
    for i in range(0, 5):
        txt.write(f"{W_List[i]}\n{cell[i]}\n\n")
    txt.close()

    with open(txt_path, 'r') as file:
        content_1 = file.read()

    print(">>>接入gpt中，请等待gpt响应........")

    content = gpt_generate(content_1)

    with open(txt_path, 'a') as file:  
        file.write("本周工作总结\n" + content + "\n")
        file.write(f"\n下一步计划 \n {cell_Next}")
    print(">>>导出周报文字版：成功！")

def gpt_generate(content):
    if api_key == 'None':
        print(">>>没有api key,不支持接入gpt，总结部分即将填入空字段")
        content = ''
        return content
    url = api_base
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json",
    }
    data = {
        "model": gpt_model,
        "messages": [
            {
                "role": "user",
                "content": f"这是我的本周工作内容：{content} 请列出三条工作总结，每一条前面标上序号，与我的格式相同。注意，请尽量简略。"
            }
        ]
    }

    response = requests.post(url, headers=headers, json=data)
    response_json = response.json()  # 将响应文本转换为JSON
    # 从响应JSON中提取content字段的值
    content = response_json["choices"][0]["message"]["content"]
    return content


main_page()

